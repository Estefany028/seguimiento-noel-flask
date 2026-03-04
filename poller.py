# poller.py
import os
from dotenv import load_dotenv
from google_client import sheets_service, drive_service
from openpyxl import load_workbook
from io import BytesIO

load_dotenv()

SPREADSHEET_ID = os.getenv("SPREADSHEET_BASE_ID")
RESP_SHEET = os.getenv("SHEET_RESPONSES_NAME", "Respuestas de formulario 1")
BASE_SHEET = os.getenv("SHEET_BASE_NAME", "Base_Personas")

def extraer_file_id(url: str):
    if not url:
        return None
    s = str(url)
    import re
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", s) or re.search(r"[?&]id=([a-zA-Z0-9_-]+)", s)
    return m.group(1) if m else None

def read_values(range_a1: str):
    svc = sheets_service()
    resp = svc.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=range_a1,
        valueRenderOption="FORMATTED_VALUE",
    ).execute()
    return resp.get("values", [])

def write_value(range_a1: str, value):
    svc = sheets_service()
    svc.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=range_a1,
        valueInputOption="USER_ENTERED",
        body={"values": [[value]]},
    ).execute()

def append_rows_to_base(rows):
    svc = sheets_service()
    svc.spreadsheets().values().append(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{BASE_SHEET}!A1",
        valueInputOption="USER_ENTERED",
        insertDataOption="INSERT_ROWS",
        body={"values": rows},
    ).execute()

def download_drive_file_bytes(file_id: str) -> bytes:
    drv = drive_service()
    req = drv.files().get_media(fileId=file_id)
    from googleapiclient.http import MediaIoBaseDownload
    import io
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    return fh.getvalue()

def parse_xlsx_from_row11(xlsx_bytes: bytes):
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # fila 11 (1-based)
    start_row = 11
    max_row = ws.max_row
    max_col = ws.max_column

    out = []
    for r in range(start_row, max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if any(v not in (None, "") for v in row):
            out.append([("" if v is None else v) for v in row])

    return out

def main():
    # 1) cursor
    cursor_vals = read_values("Config!A2")
    last_processed = int(cursor_vals[0][0]) if cursor_vals and cursor_vals[0] else 2

    # 2) leer respuestas (A:C)
    resp = read_values(f"{RESP_SHEET}!A1:C")
    if len(resp) <= last_processed:
        print("Sin nuevas solicitudes.")
        return

    nuevas = resp[last_processed:]  # last_processed es índice 1-based-like guardado
    procesadas = 0

    for fila in nuevas:
        link = fila[2] if len(fila) > 2 else ""
        file_id = extraer_file_id(link)
        if not file_id:
            last_processed += 1
            continue

        xlsx_bytes = download_drive_file_bytes(file_id)
        rows = parse_xlsx_from_row11(xlsx_bytes)

        if rows:
            append_rows_to_base(rows)
            # TODO: aquí llamas tus validaciones en Python (certificados / inducción / SS)
            procesadas += 1

        last_processed += 1

    # 3) guardar cursor
    write_value("Config!A2", last_processed)
    print(f"Procesadas: {procesadas}")

if __name__ == "__main__":
    main()
